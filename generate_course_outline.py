# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ===================== COLOR PALETTE =====================
DARK_BLUE = "1B3A5C"       # 深蓝 主色
MID_BLUE = "2C5F8A"        # 中蓝 表头
LIGHT_BLUE = "D6E6F5"      # 浅蓝 行间隔
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F7FA"      # 极浅灰 背景
ACCENT_GOLD = "D4A843"     # 金色 accent
FONT_DARK = "1A1A2E"
FONT_WHITE = "FFFFFF"

# ===================== STYLES =====================
thin_border = Border(
    left=Side(style='thin', color='B0C4DE'),
    right=Side(style='thin', color='B0C4DE'),
    top=Side(style='thin', color='B0C4DE'),
    bottom=Side(style='thin', color='B0C4DE'),
)

title_font = Font(name='微软雅黑', size=18, bold=True, color=DARK_BLUE)
subtitle_font = Font(name='微软雅黑', size=11, color=MID_BLUE)
header_font = Font(name='微软雅黑', size=11, bold=True, color=FONT_WHITE)
cell_font = Font(name='微软雅黑', size=10, color=FONT_DARK)
module_font = Font(name='微软雅黑', size=10, bold=True, color=DARK_BLUE)
note_font = Font(name='微软雅黑', size=9, color='555555')
gold_font = Font(name='微软雅黑', size=10, bold=True, color=ACCENT_GOLD)

header_fill = PatternFill(start_color=MID_BLUE, end_color=MID_BLUE, fill_type='solid')
alt_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type='solid')
white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type='solid')

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def apply_row(ws, row, data, fill=None, font=None, alignment=None):
    for col_idx, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        cell.border = thin_border

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ===================== SHEET 1: 班课大纲 =====================
ws1 = wb.active
ws1.title = "班课大纲（10课时）"
ws1.sheet_properties.tabColor = MID_BLUE

# Title
ws1.merge_cells('A1:F1')
ws1['A1'] = "《数学建模竞赛获奖冲刺》课程规划表（班课·10课时）"
ws1['A1'].font = title_font
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 45

# Subtitle
ws1.merge_cells('A2:F2')
ws1['A2'] = "适用场景：大班教学（20-40人）  |  授课目标：10课时内具备国赛/美赛获奖能力  |  版本：V1.0"
ws1['A2'].font = subtitle_font
ws1['A2'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[2].height = 28

# Header
headers = ['课时', '课程模块', '课程名称', '主要内容', '时长', '教学形式']
for col_idx, h in enumerate(headers, 1):
    cell = ws1.cell(row=3, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
ws1.row_dimensions[3].height = 30

# Data
group_data = [
    ['1', '背景引入', '数学建模竞赛全景概览',
     '国赛/美赛/研赛等主流赛事特点与对比；评奖规则与评分标准解读；参赛流程与时间线；队员分工与角色定位（建模手/编程手/写作手）；获奖团队经验要点',
     '1h', '讲授+案例'],
    ['2', '工具应用', 'MATLAB与Python编程实战',
     'MATLAB矩阵运算、绘图、常用工具箱；Python科学计算栈（NumPy/Pandas/SciPy/Matplotlib）；数据读取与预处理；代码模板与调试技巧',
     '1h', '讲授+实操'],
    ['3', '工具应用', 'AI辅助建模实战',
     'Claude与GPT在数模中的高阶用法（思路启发、代码生成、论文润色）；数模Agent工具链应用；AI辅助下的工作流优化；注意事项与常见陷阱',
     '1h', '讲授+演示'],
    ['4', '基础模型', '优化模型与评价模型',
     '线性/非线性规划、整数规划、多目标规划；层次分析法（AHP）、模糊综合评价、TOPSIS；模型选择方法论——如何根据问题特征快速匹配模型',
     '1h', '讲授+案例'],
    ['5', '基础模型', '预测与统计分析模型',
     '回归分析、时间序列预测（ARIMA）、灰色预测GM(1,1)；插值与拟合；聚类分析（K-means）、主成分分析（PCA）；灵敏度分析与模型检验',
     '1h', '讲授+案例'],
    ['6', '基础模型', '微分方程与机理建模',
     '常微分方程建模（人口/传染/物理问题）；差分方程模型；元胞自动机；机理建模与数据驱动建模的融合思路',
     '1h', '讲授+案例'],
    ['7', '基础模型', '图论与网络模型',
     '图论基础（最短路、最小生成树、TSP）；排队论；蒙特卡洛模拟；模型判断能力专项训练——给定问题快速定位模型框架',
     '1h', '讲授+训练'],
    ['8', '论文写作', '论文写作规范与技巧',
     '论文结构拆解（摘要→问题分析→模型建立→求解→检验→评价）；摘要写作魔鬼训练（决定获奖下限的关键）；图表规范与可视化技巧；常用学术表达句式',
     '1h', '讲授+练习'],
    ['9', '论文写作', '优秀论文解析与模拟写作',
     '历年国赛/美赛优秀论文精读与分析；常见写作扣分点避坑；限时写作训练（1h内完成摘要+模型概述）；同伴互评与修改',
     '1h', '研讨+实操'],
    ['10', '综合训练', '全真模拟与赛前冲刺',
     '限时全真模拟（抽练往年赛题）；完整流程走一遍（选题→建模→编程→写作）；时间分配策略与应急预案；常见翻车场景及应对方案；课程总结与赛前checklist',
     '1h', '模拟+点评'],
]

for r_idx, row_data in enumerate(group_data, 4):
    fill = alt_fill if (r_idx % 2 == 0) else white_fill
    apply_row(ws1, r_idx, row_data, fill=fill, font=cell_font, alignment=left_align)
    ws1.row_dimensions[r_idx].height = 60
    # Center the first 3 columns
    for c in [1, 2, 5, 6]:
        ws1.cell(row=r_idx, column=c).alignment = center_align

# Module column formatting
module_colors = {
    '背景引入': '5B8DB8',
    '工具应用': '4A7C9E',
    '基础模型': '3A6B8E',
    '论文写作': '6B9EC0',
    '综合训练': '8BB5D8',
}
for row_idx in range(4, 14):
    module = ws1.cell(row=row_idx, column=2).value
    if module in module_colors:
        ws1.cell(row=row_idx, column=2).font = Font(
            name='微软雅黑', size=10, bold=True,
            color=module_colors[module]
        )

# Summary row
summary_row = 14
ws1.merge_cells(f'A{summary_row}:C{summary_row}')
ws1.cell(row=summary_row, column=1, value='合计').font = Font(name='微软雅黑', size=11, bold=True, color=DARK_BLUE)
ws1.cell(row=summary_row, column=1).alignment = center_align
ws1.cell(row=summary_row, column=1).border = thin_border
ws1.cell(row=summary_row, column=2).border = thin_border
ws1.cell(row=summary_row, column=3).border = thin_border
ws1.merge_cells(f'D{summary_row}:E{summary_row}')
ws1.cell(row=summary_row, column=4, value='总课时：10课时  |  总授课时长：10小时  |  覆盖模型类型：6大类').font = Font(name='微软雅黑', size=10, bold=True, color=DARK_BLUE)
ws1.cell(row=summary_row, column=4).alignment = left_align
ws1.cell(row=summary_row, column=4).border = thin_border
ws1.cell(row=summary_row, column=5).border = thin_border
ws1.cell(row=summary_row, column=6, value='10h').font = Font(name='微软雅黑', size=10, bold=True, color=MID_BLUE)
ws1.cell(row=summary_row, column=6).alignment = center_align
ws1.cell(row=summary_row, column=6).border = thin_border

# Column widths
set_col_widths(ws1, [6, 12, 22, 58, 8, 14])

# ===================== SHEET 2: 一对一课程大纲 =====================
ws2 = wb.create_sheet("一对一课程（10课时）")
ws2.sheet_properties.tabColor = ACCENT_GOLD

# Title
ws2.merge_cells('A1:F1')
ws2['A1'] = "《数学建模竞赛获奖冲刺》课程规划表（一对一·10课时）"
ws2['A1'].font = title_font
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 45

# Subtitle
ws2.merge_cells('A2:F2')
ws2['A2'] = "适用场景：一对一私教定制  |  特色：团队启动 + 赛中指导（2课时）  |  版本：V1.0"
ws2['A2'].font = subtitle_font
ws2['A2'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[2].height = 28

# Headers
for col_idx, h in enumerate(headers, 1):
    cell = ws2.cell(row=3, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color=ACCENT_GOLD, end_color=ACCENT_GOLD, fill_type='solid')
    cell.alignment = center_align
    cell.border = thin_border
ws2.row_dimensions[3].height = 30

one_on_one_data = [
    ['1', '团队启动+背景引入', '队伍组建与竞赛全景',
     '① 队伍组建：队员角色定位与分工、队员能力评估与补位策略、学习规划与竞赛时间线制定\n② 竞赛概览：国赛/美赛/研赛对比、评分标准、获奖路径拆解',
     '1h', '面谈+规划'],
    ['2', '工具应用', 'MATLAB编程实战',
     'MATLAB矩阵运算/绘图/工具箱；常用算法模板库建立（优化求解/数据拟合/统计计算）；代码调试与性能优化；一对一针对性补齐编程短板',
     '1h', '讲授+实操'],
    ['3', '工具应用', 'Python与AI辅助建模',
     'Python科学计算栈（NumPy/Pandas/Scikit-learn）；Claude与GPT在数模中的高阶用法；数模Agent工具链；AI辅助下的工作流提效；工具选型策略',
     '1h', '讲授+演示'],
    ['4', '基础模型', '优化·评价·预测模型',
     '线性/非线性规划、整数规划、多目标规划；AHP/TOPSIS/模糊综合评价；回归分析、时间序列、灰色预测；★ 模型判断力训练：给定赛题快速匹配模型框架',
     '1h', '精讲+训练'],
    ['5', '基础模型', '微分方程·图论·机理建模',
     '常微分方程建模（人口/物理/传染模型）；图论基础（最短路/最小生成树/TSP）；元胞自动机与蒙特卡洛模拟；★ 综合模型选择实战演练',
     '1h', '精讲+训练'],
    ['6', '论文写作', '论文写作精讲',
     '论文结构拆解（摘要→问题分析→建模→求解→检验→评价）；摘要写作特训（决定获奖下限的关键）；图表规范与可视化；常见扣分点及避坑指南',
     '1h', '讲授+练习'],
    ['7', '论文写作', '优秀论文解析与写作实战',
     '历年优秀论文精读：结构模仿与创新点分析；限时写作训练（1h内完成摘要+模型框架）；一对一逐段批改反馈',
     '1h', '研讨+批改'],
    ['8', '综合训练', '全真模拟与赛前冲刺',
     '限时全真模拟赛题演练；选题→建模→编程→写作全流程计时跑通；时间分配策略与应急预案；赛前checklist',
     '1h', '模拟+点评'],
    ['9', '赛中指导', '模型确认与思路纠偏',
     '赛中实时支持：选题方向确认把关、建模思路合理性校验、模型选择纠偏、代码实现卡点排查；确保前中期不跑偏',
     '比赛期间', '在线指导'],
    ['10', '赛中指导', '论文检查与最终冲刺',
     '赛中实时支持：论文结构完整性检查、摘要/关键词打磨、图表规范化、灵敏度分析检验、常见遗漏项排查、最终提交前复核',
     '比赛期间', '在线指导'],
]

for r_idx, row_data in enumerate(one_on_one_data, 4):
    fill = alt_fill if (r_idx % 2 == 0) else white_fill
    # For 赛中指导 rows, use a warmer fill
    if '赛中指导' in row_data[1]:
        fill = PatternFill(start_color='FFF8E8', end_color='FFF8E8', fill_type='solid')
    apply_row(ws2, r_idx, row_data, fill=fill, font=cell_font, alignment=left_align)
    ws2.row_dimensions[r_idx].height = 70
    for c in [1, 2, 5, 6]:
        ws2.cell(row=r_idx, column=c).alignment = center_align

# Module color formatting - second sheet
module_colors2 = {
    '团队启动+背景引入': '2C5F8A',
    '工具应用': '4A7C9E',
    '基础模型': '3A6B8E',
    '论文写作': '6B9EC0',
    '综合训练': '8BB5D8',
    '赛中指导': 'D4A843',
}
for row_idx in range(4, 14):
    module = ws2.cell(row=row_idx, column=2).value
    if module in module_colors2:
        ws2.cell(row=row_idx, column=2).font = Font(
            name='微软雅黑', size=10, bold=True,
            color=module_colors2[module]
        )

# Summary row
sr = 14
ws2.merge_cells(f'A{sr}:C{sr}')
ws2.cell(row=sr, column=1, value='合计').font = Font(name='微软雅黑', size=11, bold=True, color=DARK_BLUE)
ws2.cell(row=sr, column=1).alignment = center_align
ws2.cell(row=sr, column=1).border = thin_border
ws2.cell(row=sr, column=2).border = thin_border
ws2.cell(row=sr, column=3).border = thin_border
ws2.merge_cells(f'D{sr}:E{sr}')
ws2.cell(row=sr, column=4, value='总课时：10课时（含2课时赛中指导）  |  赛前完成8课时闭环  |  赛中实时护航').font = Font(name='微软雅黑', size=10, bold=True, color=DARK_BLUE)
ws2.cell(row=sr, column=4).alignment = left_align
ws2.cell(row=sr, column=4).border = thin_border
ws2.cell(row=sr, column=5).border = thin_border
ws2.cell(row=sr, column=6, value='10h+赛中').font = Font(name='微软雅黑', size=10, bold=True, color=ACCENT_GOLD)
ws2.cell(row=sr, column=6).alignment = center_align
ws2.cell(row=sr, column=6).border = thin_border

set_col_widths(ws2, [6, 16, 22, 58, 10, 12])

# ===================== SHEET 3: 对比总览 =====================
ws3 = wb.create_sheet("方案对比总览")
ws3.sheet_properties.tabColor = DARK_BLUE

ws3.merge_cells('A1:E1')
ws3['A1'] = "班课 vs 一对一方案对比总览"
ws3['A1'].font = Font(name='微软雅黑', size=16, bold=True, color=DARK_BLUE)
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 40

comp_headers = ['对比维度', '班课（10课时）', '一对一（10课时）', '一对一调整说明', '建议选择场景']
comp_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
for col_idx, h in enumerate(comp_headers, 1):
    cell = ws3.cell(row=2, column=col_idx, value=h)
    cell.font = Font(name='微软雅黑', size=10, bold=True, color=FONT_WHITE)
    cell.fill = comp_fill
    cell.alignment = center_align
    cell.border = thin_border
ws3.row_dimensions[2].height = 30

comp_data = [
    ['团队建设', '无单独模块', '第1课时启动（0.5h）', '新增团队搭建与规划环节', '有固定队伍需磨合的学员'],
    ['基础模型', '4课时（全面展开）', '2课时（压缩精讲）', '压减模型通讲，强化判断力训练', '有一定基础或时间紧迫者'],
    ['论文写作', '2课时（讲授+互评）', '2课时（讲授+批改）', '由同伴互评升级为一对一批改', '论文是短板需深度打磨者'],
    ['赛中指导', '无', '2课时（赛中实时支持）', '新增赛中模型纠偏+论文检查', '冲击高奖项需要全程护航者'],
    ['编程工具', '2课时（MATLAB+Python+AI）', '2课时（MATLAB+Python+AI）', '可根据学员短板动态调整侧重', '编程基础参差不齐均可'],
    ['综合模拟', '1课时限时模拟', '1课时全真模拟+赛前冲刺', '模拟后一对一复盘点评', '需要赛前状态调整者'],
    ['总课时', '10课时=10h', '10课时=8h+赛中2h', '赛中指导计入课时但提高转化率', '—'],
    ['收费定位', '标准价（人均低）', '定制价（单价高）', '—', '—'],
    ['目标产出', '具备获奖能力', '冲击国一/国奖/美赛F/O奖', '赛中指导提升获奖确定性', '—'],
]

for r_idx, row_data in enumerate(comp_data, 3):
    fill = alt_fill if (r_idx % 2 == 1) else white_fill
    for col_idx, val in enumerate(row_data, 1):
        cell = ws3.cell(row=r_idx, column=col_idx, value=val)
        cell.font = cell_font
        cell.fill = fill
        cell.alignment = left_align if col_idx == len(row_data) else center_align
        cell.border = thin_border
    ws3.row_dimensions[r_idx].height = 35

# Recommendation section
rec_row = 3 + len(comp_data) + 1
ws3.merge_cells(f'A{rec_row}:E{rec_row}')
ws3.cell(row=rec_row, column=1, value='★ 推荐策略').font = Font(name='微软雅黑', size=12, bold=True, color=DARK_BLUE)
ws3.cell(row=rec_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
ws3.row_dimensions[rec_row].height = 30

recs = [
    '• 预算有限、参赛经验较少的学员群体 → 推荐班课方案，性价比高，系统覆盖所有必要知识点',
    '• 有一定基础、目标冲击高奖项的学员 → 推荐一对一方案，赛中指导是关键变量，能有效降低临场失误率',
    '• 混合模式（建议）：班课讲授基础内容 + 赛中单独购买一对一指导，兼顾成本与效果',
]
for i, rec in enumerate(recs):
    r = rec_row + 1 + i
    ws3.merge_cells(f'A{r}:E{r}')
    ws3.cell(row=r, column=1, value=rec).font = Font(name='微软雅黑', size=10, color='333333')
    ws3.cell(row=r, column=1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws3.row_dimensions[r].height = 28

set_col_widths(ws3, [14, 18, 22, 28, 22])

# ===================== SAVE =====================
output_path = r'C:\Users\24404\bzd-website\数学建模竞赛课程规划方案.xlsx'
wb.save(output_path)
print(f"文件已保存至：{output_path}")
