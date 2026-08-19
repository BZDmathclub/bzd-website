import openpyxl
import json

SRC = r'C:\Users\24404\Documents\Codex\2026-08-19\file-c-users-24404-desktop-2025\outputs\高教社杯国赛学校综合统计与2026预测.xlsx'
OUT = r'C:\Users\24404\bzd-website\data\national-awards.js'

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb.active

# 列结构: 1名称 2赛区 3-4(21年一二) 5(21合计) 6-7(22一二) 8(22合计) 9-10(23一二) 11(23合计)
# 12-13(24一二) 14(24合计) 15-16(25一二) 17(25合计) 18历年一等 19历年二等 20历年合计 21预测 22指导老师 23次数
rows = []
years = [2021, 2022, 2023, 2024, 2025]
for r in range(2, ws.max_row + 1):
    name = ws.cell(row=r, column=1).value
    if not name:
        continue
    region = ws.cell(row=r, column=2).value or ''
    yearly = []
    for i in range(5):
        col_a = 3 + i * 3
        col_b = 4 + i * 3
        a = ws.cell(row=r, column=col_a).value or 0
        b = ws.cell(row=r, column=col_b).value or 0
        yearly.append([int(a), int(b)])
    total = ws.cell(row=r, column=20).value or 0
    pred = ws.cell(row=r, column=21).value or 0
    teacher = ws.cell(row=r, column=22).value or ''
    count = ws.cell(row=r, column=23).value or 0
    rows.append({
        'n': str(name).strip(),
        'rg': str(region).strip(),
        'y': yearly,
        't': int(total),
        'p': int(pred),
        'tc': str(teacher).strip(),
        'c': int(count),
    })

js = 'window.NATIONAL_AWARDS_DATA = ' + json.dumps(rows, ensure_ascii=False) + ';'
with open(OUT, 'w', encoding='utf-8') as f:
    f.write(js)

print(f'共 {len(rows)} 所高校')
print(f'输出文件大小: {len(js.encode("utf-8"))} bytes')
